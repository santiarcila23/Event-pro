"""
utils/exportar.py
Exportación a Excel (openpyxl) y PDF (reportlab)
"""
import os
from datetime import datetime

# ── Excel ────────────────────────────────────────────────────────
def exportar_excel(titulo: str, encabezados: list, filas: list, ruta: str = None) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not ruta:
        ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta = os.path.join("exports", f"{titulo.replace(' ','_')}_{ts}.xlsx")
    os.makedirs(os.path.dirname(ruta) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    # Título principal
    ws.merge_cells(f"A1:{chr(64+len(encabezados))}1")
    cell_titulo = ws["A1"]
    cell_titulo.value = titulo
    cell_titulo.font      = Font(bold=True, size=14, color="FFFFFF")
    cell_titulo.fill      = PatternFill("solid", fgColor="1B4F72")
    cell_titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Encabezados
    header_fill = PatternFill("solid", fgColor="2E86AB")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col, enc in enumerate(encabezados, 1):
        c = ws.cell(row=2, column=col, value=enc)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border    = border

    # Datos
    for i, fila in enumerate(filas):
        fill = PatternFill("solid", fgColor="EBF5FB") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col, valor in enumerate(fila, 1):
            c = ws.cell(row=i+3, column=col, value=str(valor) if valor is not None else "")
            c.fill   = fill
            c.border = border
            c.alignment = Alignment(horizontal="left")

    # Auto-ancho
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Fecha de exportación
    ws.cell(row=len(filas)+4, column=1, value=f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    wb.save(ruta)
    return ruta


# ── PDF ──────────────────────────────────────────────────────────
def exportar_pdf(titulo: str, encabezados: list, filas: list, ruta: str = None) -> str:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    if not ruta:
        ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta = os.path.join("exports", f"{titulo.replace(' ','_')}_{ts}.pdf")
    os.makedirs(os.path.dirname(ruta) or ".", exist_ok=True)

    doc = SimpleDocTemplate(ruta, pagesize=landscape(letter),
                             leftMargin=0.5*inch, rightMargin=0.5*inch,
                             topMargin=0.5*inch, bottomMargin=0.5*inch)

    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle("titulo", parent=styles["Title"],
                                  fontSize=16, textColor=colors.HexColor("#1B4F72"),
                                  spaceAfter=12)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                                fontSize=9, textColor=colors.grey, spaceAfter=6)

    story = [
        Paragraph(titulo, titulo_style),
        Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total registros: {len(filas)}", sub_style),
        Spacer(1, 10),
    ]

    # Tabla
    data = [encabezados] + [[str(v) if v is not None else "" for v in fila] for fila in filas]
    col_w = (9.5 * inch) / len(encabezados)
    tabla = Table(data, colWidths=[col_w] * len(encabezados), repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0),  colors.HexColor("#1B4F72")),
        ("TEXTCOLOR",   (0,0), (-1,0),  colors.white),
        ("FONTNAME",    (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0),  10),
        ("ALIGN",       (0,0), (-1,0),  "CENTER"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#EBF5FB"), colors.white]),
        ("FONTSIZE",    (0,1), (-1,-1), 9),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("PADDING",     (0,0), (-1,-1), 4),
    ]))
    story.append(tabla)
    doc.build(story)
    return ruta
